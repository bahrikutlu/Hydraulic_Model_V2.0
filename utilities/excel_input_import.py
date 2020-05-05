import openpyxl
from definitions import input_well


class ImportFromExcel:
    def __init__(self, sheet_name):
        # opens the input xlsx and loads the sheet name that is provided. sheet name is provided as string
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(input_well)
        self.sheet = self.wb[sheet_name]
        self.drill_string = list()
        self.bottom_hole_assembly = list()
        self.casing_design = list()
        self.bit_nozzles = list()
        self.rheology = list()
        self.string = list()
        self.bit_data = list()

    def all_casing_strings(self):
        for column in range(2, 12):
            for row in range(2, 10):
                cell = self.sheet.cell(row, column)
                if cell.value is not None:
                    self.string.append(cell.value)
            if self.string:
                self.casing_design.append(self.string)
            self.string = list()
        return self.casing_design

    def all_drill_strings(self):
        for column in range(2, 12):
            for row in range(12, 20):
                cell = self.sheet.cell(row, column)
                if cell.value is not None:
                    self.string.append(cell.value)
            if self.string:
                self.drill_string.append(self.string)
            self.string = list()
        return self.drill_string

    def all_bha(self):
        for column in range(2, 12):
            for row in range(22, 30):
                cell = self.sheet.cell(row, column)
                if cell.value is not None:
                    self.string.append(cell.value)
            if self.string:
                self.bottom_hole_assembly.append(self.string)
            self.string = list()
        return self.bottom_hole_assembly

    def bit(self):
        for row in range(32, 36):
            cell = self.sheet.cell(row, 2)
            if cell.value is not None:
                self.string.append(cell.value)
        if self.string:
            self.bit_data.append(self.string)
        self.string = list()
        return self.bit_data

    def all_bit_nozzles(self):
        for column in range(2, 12):
            cell = self.sheet.cell(38, column)
            if cell.value is not None:
                self.string.append(cell.value)
        if self.string:
            self.bit_nozzles.append(self.string)
        self.string = list()
        return self.bit_nozzles[0]

    def hole_size(self):
        cell = self.sheet.cell(56, 2)
        return cell.value

    def viscometer_readings(self):
        for column in range(1, 3):
            for row in range(46, 52):
                cell = self.sheet.cell(row, column)
                if cell.value is not None:
                    self.string.append(cell.value)
            if self.string:
                self.rheology.append(self.string)
            self.string = list()
        result = map(lambda x: x * 1.70, self.rheology[0])
        self.rheology[0] = list(result)
        return self.rheology

    def fluid_type(self):
        cell = self.sheet.cell(52, 2)
        return cell.value

    def flow_rate(self):
        cell = self.sheet.cell(53, 2)
        return cell.value

    def mud_weight(self):
        cell = self.sheet.cell(54, 2)
        return cell.value

    def eccentricity(self):
        cell = self.sheet.cell(55, 2)
        return cell.value

    def units(self):
        cell = self.sheet.cell(60, 2)
        return cell.value

    def calculation_step_difference(self):
        cell = self.sheet.cell(61, 2)
        return cell.value